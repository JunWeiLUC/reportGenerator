from openpyxl import load_workbook
from openpyxl import *
from openpyxl.styles import *
import datetime
from openpyxl.drawing.image import Image


wb = load_workbook(filename = '1.xlsx')

sheet = wb['Sheet1']

img = Image('logo2.png')
sheet.add_image(img, 'A1')

patientID = '8703163272711'

c = sheet.cell(row=6, column=7)
c.value = patientID

#读取实验结果
result = load_workbook(filename = (patientID + '.xlsx'))
resultSheet = result['Sheet1']

#读取病人信息
sampleInformation = load_workbook(filename = '样本信息登记表_浙江医院.xlsx')
sampleInformationSheet = sampleInformation ['Sheet1']

#写入病人信息
row_count = sampleInformationSheet.max_row
for n in range (1, row_count+1):
    temp = sampleInformationSheet.cell(row =n, column = 2).value
    if (temp == patientID):
        姓名 = sampleInformationSheet.cell (row = n, column = 15).value
        c = sheet.cell(row=9, column=2)
        c.value = 姓名
        性别 = sampleInformationSheet.cell(row=n, column=16).value
        c = sheet.cell(row=9, column=4)
        c.value = 性别
        出生年月 = sampleInformationSheet.cell(row=n, column=18).value
        c = sheet.cell(row=9, column=6)
        c.value = 出生年月
        联系方式 = sampleInformationSheet.cell (row = n, column = 20).value
        c = sheet.cell(row=9, column=8)
        c.value = 联系方式
        医院名 = sampleInformationSheet.cell (row = n, column = 10).value
        c = sheet.cell(row=10, column=2)
        c.value = 医院名
        住院号 = sampleInformationSheet.cell (row = n, column = 11).value
        c = sheet.cell(row=10, column=4)
        c.value = 住院号
        床号 = sampleInformationSheet.cell (row = n, column = 12).value
        c = sheet.cell(row=10, column=6)
        c.value = 床号
        入院时间 = sampleInformationSheet.cell (row = n, column = 13).value
        c = sheet.cell(row=10, column=8)
        c.value = 入院时间
        送检日期 = sampleInformationSheet.cell (row = n, column = 6).value
        c = sheet.cell(row=11, column=4)
        c.value = 送检日期
        送检医师 = sampleInformationSheet.cell (row = n, column = 9).value
        c = sheet.cell(row=11, column=6)
        c.value = 送检医师
        门诊号 = sampleInformationSheet.cell (row = n, column = 14).value
        c = sheet.cell(row=11, column=8)
        c.value = 门诊号
        临床诊断 = sampleInformationSheet.cell(row=n, column=21).value
        c = sheet.cell(row=12, column=2)
        c.value = 临床诊断



#读取结果分析数据库
dataInterpretation = load_workbook (filename = 'CVD Pharmagenetics database_04.xlsx')
dataSheet = dataInterpretation['RS']






#查找华法林实验结果，写入报告
snpNumber = 5
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '华法林'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找氯吡格雷实验结果，写入报告
snpNumber = 3
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '氯吡格雷'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找阿司匹林实验结果，写入报告
snpNumber = 9
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '阿司匹林'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找CCB类：氨氯地平、硝苯地平、维拉帕米等实验结果，写入报告
snpNumber = 1
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == 'CCB类：氨氯地平、硝苯地平、维拉帕米等'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''


#查找ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等实验结果，写入报告
snpNumber = 2
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == 'ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''


#查找BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等实验结果，写入报告
snpNumber = 2
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == 'BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找ACEI类：卡托普利、依那普利、苯那普利、赖诺普利、培哚普利、福辛普利等实验结果，写入报告
snpNumber = 7
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == 'ACEI类：卡托普利、依那普利、苯那普利、赖诺普利、培哚普利、福辛普利等'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找利尿剂：氢氯噻嗪、布美他尼、呋塞米、托拉塞米等实验结果，写入报告
snpNumber = 1
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '利尿剂：氢氯噻嗪、布美他尼、呋塞米、托拉塞米等'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''

#查找他汀类实验结果，写入报告
snpNumber = 6
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '他汀类'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''


#查找硝酸甘油实验结果，写入报告
snpNumber = 1
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '硝酸甘油'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''


#查找单硝酸异山梨酯实验结果，写入报告
snpNumber = 2
row_count = sheet.max_row
for n in range (1, row_count+1):
    temp = sheet.cell(row =n, column = 1).value
    if (temp == '单硝酸异山梨酯'):
        snpStart = n
        print (snpStart)
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            for i in range (1,50):
                rs2 = resultSheet.cell (row = i, column = 4).value
                if (rs1 == rs2):
                    检测结果 = resultSheet.cell (row = i, column = 5).value
                    c = sheet.cell (row = snpStart + j, column = 4)
                    c.value = 检测结果
                    break

        #在结果分析数据库查找结果的分析，写入报告
        for j in range (0,snpNumber):
            rs1 = sheet.cell (row = snpStart + j, column = 3).value
            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
            rsr1r = rsr1 [::-1]
            for i in range (1,117):
                rs2 = dataSheet.cell (row = i, column = 7).value
                rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2)):
                    基因型 = dataSheet.cell (row = i, column = 11).value
                    结果说明 = dataSheet.cell(row=i, column=12).value
                    c1 = sheet.cell (row = snpStart + j, column = 5)
                    c1.value = 基因型
                    c2 = sheet.cell(row=snpStart + j, column=6)
                    c2.value = 结果说明
                    结果解释 = dataSheet.cell (row = i, column = 13).value
                    c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                    基因名 = sheet.cell (row = snpStart + j, column = 2).value
                    c3.value = 基因名 + '的' +  rs1 + '位点的' + 结果解释
                    break

        for j in range (0,snpNumber):
            c = sheet.cell(row = snpStart + j, column = 5)
            rs = c.value
            if(rs == None):
                c.value = '错误'
                c.font = Font(color=colors.RED)


        '''
            需加入代谢酶型
       '''


row_count = sheet.max_row
ErrorNumber = 0
summaryReport = load_workbook(filename = '生成报告总结.xlsx')
summarySheet = summaryReport['Sheet1']
reportNumber = summarySheet.max_row
c = summarySheet.cell(row = reportNumber+1, column = 1)
c.value = patientID
experimentalID = resultSheet.cell (row =2,column=2).value
c = summarySheet.cell(row = reportNumber+1, column = 2)
c.value = experimentalID
c = summarySheet.cell(row = reportNumber+1, column = 3)
c.value = datetime.datetime.now()
for n in range (1, row_count+1):
    temp = str (sheet.cell(row =n, column = 5).value)
    if (temp == '错误'):
        ErrorNumber += 1
    c = summarySheet.cell(row=reportNumber + 1, column=4)
    c.value = ErrorNumber







wb.save (filename = (patientID +'_report.xlsx'))
summaryReport.save (filename = '生成报告总结.xlsx')

