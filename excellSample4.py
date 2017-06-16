from openpyxl import load_workbook
from openpyxl import *
from openpyxl.styles import *
import datetime
from openpyxl.drawing.image import Image


def checkExist(reportNumber, patientID):
    reportNo = 0
    for n in range (1, reportNumber+1):
        reportID = str(summarySheet.cell(row=n,column=1).value)
        if (reportID == patientID):
            reportNo = n
            break

    return reportNo



def generateReport():
   
    wb = load_workbook(filename = '1.xlsx')

    sheet = wb['Sheet1']

    img = Image('logo3.png')
    sheet.add_image(img, 'A1')



    c = sheet.cell(row=6, column=7)
    c.value = '样品编号： ' + patientID

    #读取实验结果
    result = load_workbook(filename = (patientID + '.xlsx'))
    resultSheet = result['Sheet1']

    #读取病人信息
    sampleInformation = load_workbook(filename = '样本信息登记表_浙江医院.xlsx')
    sampleInformationSheet = sampleInformation ['Sheet1']

    #写入病人信息
    row_count = sampleInformationSheet.max_row
    for n in range (1, row_count+1):
            temp = sampleInformationSheet.cell(row =n, column = 2).value
            if (temp == patientID):
                    姓名 = sampleInformationSheet.cell (row = n, column = 15).value
                    c = sheet.cell(row=9, column=2)
                    c.value = 姓名
                    性别 = sampleInformationSheet.cell(row=n, column=16).value
                    c = sheet.cell(row=9, column=4)
                    c.value = 性别
                    出生年月 = sampleInformationSheet.cell(row=n, column=18).value
                    c = sheet.cell(row=9, column=6)
                    c.value = 出生年月
                    联系方式 = sampleInformationSheet.cell (row = n, column = 20).value
                    c = sheet.cell(row=9, column=8)
                    c.value = 联系方式
                    医院名 = sampleInformationSheet.cell (row = n, column = 10).value
                    c = sheet.cell(row=10, column=2)
                    c.value = 医院名
                    住院号 = sampleInformationSheet.cell (row = n, column = 11).value
                    c = sheet.cell(row=10, column=4)
                    c.value = 住院号
                    床号 = sampleInformationSheet.cell (row = n, column = 12).value
                    c = sheet.cell(row=10, column=6)
                    c.value = 床号
                    入院时间 = sampleInformationSheet.cell (row = n, column = 13).value
                    c = sheet.cell(row=10, column=8)
                    c.value = 入院时间
                    送检日期 = sampleInformationSheet.cell (row = n, column = 6).value
                    c = sheet.cell(row=11, column=4)
                    c.value = 送检日期
                    送检医师 = sampleInformationSheet.cell (row = n, column = 9).value
                    c = sheet.cell(row=11, column=6)
                    c.value = 送检医师
                    门诊号 = sampleInformationSheet.cell (row = n, column = 14).value
                    c = sheet.cell(row=11, column=8)
                    c.value = 门诊号
                    临床诊断 = sampleInformationSheet.cell(row=n, column=21).value
                    c = sheet.cell(row=12, column=2)
                    c.value = 临床诊断



    #读取结果分析数据库
    dataInterpretation = load_workbook (filename = 'CVD Pharmagenetics database_04.xlsx')
    dataSheet = dataInterpretation['RS']
    enzyGenoSheet = dataInterpretation['ENZYCALL']
    enzyGenoIPSheet = dataInterpretation['代谢酶型']


    #定义代谢酶型 CYP2C9
    def CYP2C9(SNP1, SNP2, sheet1):
            result = 'ERROR'
            callNumber = sheet1.max_row
            for n in range (1,callNumber+1):
                    temp = str(sheet1.cell(row=n,column=1).value)
                    if(temp == 'CYP2C9'):
                            for i in range (0,6):
                                    ez1 = str(sheet1.cell(row=n+1+i, column=2).value)
                                    ez2 = str(sheet1.cell(row=n+1+i, column=3).value)
                                    if (SNP1 == ez1 and SNP2 == ez2):
                                            result = str(sheet1.cell(row=n+1+i, column=1).value)
                                            break

            print(result)
            return result


    #定义代谢酶型 CYP2C19
    def CYP2C19(SNP1, SNP2, SNP3, sheet1):
            result = 'ERROR'
            callNumber = sheet1.max_row
            for n in range (1,callNumber+1):
                    temp = str(sheet1.cell(row=n,column=1).value)
                    if(temp == 'CYP2C19'):
                            for i in range (0,10):
                                    ez1 = str(sheet1.cell(row=n+1+i, column=2).value)
                                    ez2 = str(sheet1.cell(row=n+1+i, column=3).value)
                                    ez3 = str(sheet1.cell(row=n+1+i, column=4).value)
                                    if (SNP1 == ez1 and SNP2 == ez2 and SNP3 == ez3):
                                            result = str(sheet1.cell(row=n+1+i, column=1).value)
                                            break

            print(result)
            return result


    #定义代谢酶型 CYP2D6
    def CYP2D6(SNP1, SNP2, SNP3, sheet1):
            result = 'ERROR'
            callNumber = sheet1.max_row
            for n in range (1,callNumber+1):
                    temp = str(sheet1.cell(row=n,column=1).value)
                    if(temp == 'CYP2D6'):
                            for i in range (0,10):
                                    ez1 = str(sheet1.cell(row=n+1+i, column=2).value)
                                    ez2 = str(sheet1.cell(row=n+1+i, column=3).value)
                                    ez3 = str(sheet1.cell(row=n+1+i, column=4).value)
                                    if (SNP1 == ez1 and SNP2 == ez2 and SNP3 == ez3):
                                            result = str(sheet1.cell(row=n+1+i, column=1).value)
                                            break

            print(result)
            return result



    #定义代谢酶型 SLCO1B1
    def SLCO1B1(SNP1, SNP2, sheet1):
            result = 'ERROR'
            callNumber = sheet1.max_row
            for n in range (1,callNumber+1):
                    temp = str(sheet1.cell(row=n,column=1).value)
                    if(temp == 'SLCO1B1'):
                            for i in range (0,10):
                                    ez1 = str(sheet1.cell(row=n+1+i, column=2).value)
                                    ez2 = str(sheet1.cell(row=n+1+i, column=3).value)
                                    if (SNP1 == ez1 and SNP2 == ez2):
                                            result = str(sheet1.cell(row=n+1+i, column=1).value)
                                            break

            print(result)
            return result

    #定义代谢酶型 APOE
    def APOE(SNP1, SNP2, sheet1):
            result = 'ERROR'
            callNumber = sheet1.max_row
            for n in range (1,callNumber+1):
                    temp = str(sheet1.cell(row=n,column=1).value)
                    if(temp == 'APOE'):
                            for i in range (0,6):
                                    ez1 = str(sheet1.cell(row=n+1+i, column=2).value)
                                    ez2 = str(sheet1.cell(row=n+1+i, column=3).value)
                                    if (SNP1 == ez1 and SNP2 == ez2):
                                            result = str(sheet1.cell(row=n+1+i, column=1).value)
                                            break

            print(result)
            return result






    #查找华法林实验结果，写入报告
    snpNumber = 4
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '华法林'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '华法林'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 3, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    # 在结果分析数据库查找结果的分析，并对代谢酶CYP2C9分型，写入报告
                    SNP1=''
                    SNP2=''
                    for i in range(1, 50):
                            rs = str(resultSheet.cell(row=i, column=4).value)
                            if (rs == 'rs1799853'):
                                    SNP1 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs1057910'):
                                    SNP2 = str(resultSheet.cell(row=i, column=5).value)

                    enzy = CYP2C9(SNP1, SNP2, enzyGenoSheet)
                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                    c.value = enzy

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range (1,enzyAnaCount+1):
                            temp1 = str(enzyGenoIPSheet.cell(row=n,column=7).value)
                            temp2 = str(enzyGenoIPSheet.cell(row=n,column=2).value)
                            if(enzy == temp1 and temp2 == '华法林'):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber*2 + 3, column=1)
                                    c3.value = '对于CYP2C9基因，' + 结果解释
                                    break


                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)




    #查找氯吡格雷实验结果，写入报告
    snpNumber = 2
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '氯吡格雷'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            print (rsr1)
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '氯吡格雷'):
                                            print(rs1 == rs2)
                                            print(rsr1 == rsr2 or rsr1r == rsr2)
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            print(基因型)
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 3, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    # 在结果分析数据库查找结果的分析，并对代谢酶CYP2C19分型，写入报告
                    SNP1=''
                    SNP2=''
                    SNP3=''
                    for i in range(1, 50):
                            rs = str(resultSheet.cell(row=i, column=4).value)
                            if (rs == 'rs12248560'):
                                    SNP1 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs4986893'):
                                    SNP2 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs4244285'):
                                    SNP3 = str(resultSheet.cell(row=i, column=5).value)

                    enzy = CYP2C19(SNP1, SNP2, SNP3,enzyGenoSheet)
                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                    c.value = enzy

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range (1,enzyAnaCount+1):
                            temp = str(enzyGenoIPSheet.cell(row=n,column=7).value)
                            temp1 = str(enzyGenoIPSheet.cell(row=n,column=2).value)
                            if(enzy == temp and temp1 == '氯吡格雷'):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber*2 + 3, column=1)
                                    c3.value = '对于CYP2C19基因，' + 结果解释
                                    break


                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    #查找阿司匹林实验结果，写入报告
    snpNumber = 9
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '阿司匹林'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '阿司匹林'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    #查找CCB类：氨氯地平、硝苯地平、维拉帕米等实验结果，写入报告
    snpNumber = 1
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == 'CCB类：氨氯地平、硝苯地平、维拉帕米等'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == 'CCB类：氨氯地平、硝苯地平、维拉帕米等'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''


    #查找ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等实验结果，写入报告
    snpNumber = 1
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == 'ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == 'ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 3, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    # 在结果分析数据库查找结果的分析，并对代谢酶SLCO1B1分型，写入报告
                    SNP1=''
                    SNP2=''
                    for i in range(1, 50):
                            rs = str(resultSheet.cell(row=i, column=4).value)
                            if (rs == 'rs2306283'):
                                    SNP1 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs4149056'):
                                    SNP2 = str(resultSheet.cell(row=i, column=5).value)

                    enzy = SLCO1B1(SNP1, SNP2, enzyGenoSheet)
                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                    c.value = enzy

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range (1,enzyAnaCount+1):
                            temp = str(enzyGenoIPSheet.cell(row=n,column=7).value)
                            temp1 = str(enzyGenoIPSheet.cell(row=n,column=2).value)
                            if(enzy == temp and temp1 == 'ARB类：缬沙坦、洛沙坦、坎地沙坦、奥美沙坦等'):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber*2 + 3, column=1)
                                    c3.value = '对于SLCO1B1基因，' + 结果解释
                                    break


                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''


    #查找BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等实验结果，写入报告
    snpNumber = 1
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == 'BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == 'BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 3, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    # 在结果分析数据库查找结果的分析，并对代谢酶CYP2D6分型，写入报告
                    '''
                    SNP1=''
                    SNP2=''
                    SNP3=''
                    for i in range(1, 50):
                            rs = str(resultSheet.cell(row=i, column=4).value)
                            if (rs == 'rs1065852'):
                                    SNP1 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs5030865'):
                                    SNP2 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs16947'):
                                    SNP3 = str(resultSheet.cell(row=i, column=5).value)

                    enzy = CYP2D6(SNP1, SNP2, SNP3,enzyGenoSheet)
                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                    c.value = enzy

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range (1,enzyAnaCount+1):
                            temp = str(enzyGenoIPSheet.cell(row=n,column=7).value)
                            temp1 = str(enzyGenoIPSheet.cell(row=n,column=2).value)
                            if(enzy == temp and temp1 == 'BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等'):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber*2 + 3, column=1)
                                    c3.value = '对于CYP2C19基因，' + 结果解释
                                    break
                    '''
                    #直接从结果报告中读取代谢酶型结果
                    检测结果s = 'ERROR'
                    for n in range(1, 50):
                            rs2 = str(resultSheet.cell(row=n, column=4).value)
                            if (rs2 == 'ENZY'):
                                    检测结果 = resultSheet.cell(row=n, column=5).value
                                    print(检测结果)
                                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                                    c.value = 检测结果
                                    检测结果s = str(检测结果)

                                    break

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range (1,enzyAnaCount+1):
                            temp = str(enzyGenoIPSheet.cell(row=n,column=7).value)
                            temp1 = str(enzyGenoIPSheet.cell(row=n,column=2).value)
                            if(检测结果s == temp and temp1 == 'BB类:美托洛尔、阿替洛尔、比索洛尔、卡维洛尔、布新洛尔等'):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber*2 + 3, column=1)
                                    c3.value = '对于CYP2基因，' + 结果解释
                                    break


                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    #查找ACEI类：卡托普利、依那普利、苯那普利、赖诺普利、培哚普利、福辛普利等实验结果，写入报告
    snpNumber = 7
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == 'ACEI类：卡托普利、依那普利、苯那普利、赖诺普利、培哚普利、福辛普利等'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == 'ACEI类：卡托普利、依那普利、苯那普利、赖诺普利、培哚普利、福辛普利等'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    #查找利尿剂：氢氯噻嗪、布美他尼、呋塞米、托拉塞米等实验结果，写入报告
    snpNumber = 1
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '利尿剂：氢氯噻嗪、布美他尼、呋塞米、托拉塞米等'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '利尿剂：氢氯噻嗪、布美他尼、呋塞米、托拉塞米等'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    #查找他汀类实验结果，写入报告
    snpNumber = 5
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '他汀类'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '他汀类' ):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 3, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                     # 在结果分析数据库查找结果的分析，并对代谢酶APOE分型，写入报告
                    SNP1 = ''
                    SNP2 = ''
                    for i in range(1, 50):
                            rs = str(resultSheet.cell(row=i, column=4).value)
                            if (rs == 'rs429358'):
                                    SNP1 = str(resultSheet.cell(row=i, column=5).value)
                            if (rs == 'rs7412'):
                                    SNP2 = str(resultSheet.cell(row=i, column=5).value)

                    enzy = APOE(SNP1, SNP2, enzyGenoSheet)
                    c = sheet.cell(row=snpStart + snpNumber, column=4)
                    c.value = enzy

                    enzyAnaCount = enzyGenoIPSheet.max_row
                    for n in range(1, enzyAnaCount + 1):
                            temp = str(enzyGenoIPSheet.cell(row=n, column=7).value)
                            if (enzy == temp):
                                    基因型 = enzyGenoIPSheet.cell(row=n, column=8).value
                                    结果说明 = enzyGenoIPSheet.cell(row=n, column=9).value
                                    c1 = sheet.cell(row=snpStart + snpNumber, column=5)
                                    c1.value = 基因型
                                    c2 = sheet.cell(row=snpStart + snpNumber, column=6)
                                    c2.value = 结果说明
                                    结果解释 = str(enzyGenoIPSheet.cell(row=n, column=10).value)
                                    c3 = sheet.cell(row=snpStart + snpNumber * 2 + 3, column=1)
                                    c3.value = '对于APOE基因，' + 结果解释
                                    break


                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''


    #查找硝酸甘油实验结果，写入报告
    snpNumber = 1
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '硝酸甘油'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '硝酸甘油'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''


    #查找单硝酸异山梨酯实验结果，写入报告
    snpNumber = 2
    row_count = sheet.max_row
    for n in range (1, row_count+1):
            temp = sheet.cell(row =n, column = 1).value
            if (temp == '单硝酸异山梨酯'):
                    snpStart = n
                    print (snpStart)
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            for i in range (1,50):
                                    rs2 = resultSheet.cell (row = i, column = 4).value
                                    if (rs1 == rs2):
                                            检测结果 = resultSheet.cell (row = i, column = 5).value
                                            c = sheet.cell (row = snpStart + j, column = 4)
                                            c.value = 检测结果
                                            break

                    #在结果分析数据库查找结果的分析，写入报告
                    for j in range (0,snpNumber):
                            rs1 = sheet.cell (row = snpStart + j, column = 3).value
                            rsr1 = str(sheet.cell (row = snpStart + j, column = 4).value)
                            rsr1r = rsr1 [::-1]
                            for i in range (1,117):
                                    rs2 = dataSheet.cell (row = i, column = 7).value
                                    rsr2 = str(dataSheet.cell (row = i, column = 10).value)
                                    project = str(dataSheet.cell (row = i, column = 2).value)
                                    if (rs1 == rs2 and (rsr1 == rsr2 or rsr1r == rsr2) and project == '单硝酸异山梨酯'):
                                            基因型 = dataSheet.cell (row = i, column = 11).value
                                            结果说明 = dataSheet.cell(row=i, column=12).value
                                            c1 = sheet.cell (row = snpStart + j, column = 5)
                                            c1.value = 基因型
                                            c2 = sheet.cell(row=snpStart + j, column=6)
                                            c2.value = 结果说明
                                            结果解释 = dataSheet.cell (row = i, column = 13).value
                                            c3 = sheet.cell (row = snpStart + snpNumber + j + 2, column = 1)
                                            位点名 = sheet.cell (row = snpStart + j, column = 3).value
                                            c3.value = str(位点名) + '的基因型为' + str(sheet.cell (row = snpStart + j, column = 4).value) + '型。' + str(结果解释)
                                            break

                    for j in range (0,snpNumber):
                            c = sheet.cell(row = snpStart + j, column = 5)
                            rs = c.value
                            if(rs == None):
                                    c.value = '错误'
                                    c.font = Font(color=colors.RED)


                    '''
                            需加入代谢酶型
               '''

    print ('样品报告生成完毕！')
    row_count = sheet.max_row
    ErrorNumber = 0

    summaryReport = load_workbook(filename = '生成报告总结.xlsx')
    summarySheet = summaryReport['Sheet1']
    reportNumber = summarySheet.max_row


    reportCheck = checkExist(reportNumber, patientID)

    if (reportCheck == 0):
        rowNumber = reportNumber+1
    else:
        rowNumber = reportCheck


    c = summarySheet.cell(row = rowNumber, column = 1)
    c.value = patientID
    print ('样品编号：'+ patientID)
    experimentalID = resultSheet.cell (row =2,column=2).value
    c = summarySheet.cell(row = rowNumber, column = 2)
    c.value = experimentalID
    print ('实验编号：'+ experimentalID)
    c = summarySheet.cell(row = rowNumber, column = 3)
    c.value = datetime.datetime.now()
    print ('报告生成时间：'+ str(c.value))
    for n in range (1, row_count+1):
            temp = str (sheet.cell(row =n, column = 5).value)
            if (temp == '错误'or temp == 'ERROR'):
                    ErrorNumber += 1
    c = summarySheet.cell(row=rowNumber, column=4)
    c.value = ErrorNumber
    print ('该报告错误数:'+ str(ErrorNumber))

    wb.save (filename = (patientID +'_report.xlsx'))
    summaryReport.save (filename = '生成报告总结.xlsx')
    





repeatReportGenerate = True
        
while repeatReportGenerate:
    try:
        patientID = input('样品编号：')
        result = load_workbook(filename = (patientID + '.xlsx'))
        summaryReport = load_workbook(filename = '生成报告总结.xlsx')
        summarySheet = summaryReport['Sheet1']
        reportNumber = summarySheet.max_row

        reportCheck = checkExist(reportNumber, patientID)

        if (reportCheck != 0):
            print ('该样品报告已经生成过')
            print ('样品编号：'+ summarySheet.cell(row=reportCheck,column=1).value)
            print ('实验编号：'+ summarySheet.cell(row=reportCheck,column=2).value)
            print ('报告生成时间：'+ str(summarySheet.cell(row=reportCheck,column=3).value))
            print ('该报告错误数'+ str(summarySheet.cell(row=reportCheck,column=4).value))

            answer =  input('你想继续生成报告吗？新的报告将会覆盖旧的报告（y/n）:')
            while ((answer != 'y') and (answer != 'n')):
                answer = input('请重新输入y/n!')
            if (answer == 'n'):
                print('报告生成终止')
                wait = input ('按回车键结束')
                RepeatGenerate = False
            else:
                generateReport()

        else:
            print('开始生成报告。')
            generateReport()

        repeat =  input('开始生成下一个报告（y/n):')
        while ((repeat != 'y') and (repeat != 'n')):
            repeat = input('请重新输入y/n!')
        if (repeat == 'n'):
            repeatReportGenerate = False
            
        
    except OSError:
        print ('无法找到样品编号。请重新输入！')




